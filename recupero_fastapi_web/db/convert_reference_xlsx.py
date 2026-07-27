"""Converte /mnt/user-data/uploads/recupero_materie_pippo.xlsx (fornito come
nuovo database di riferimento) in db/02_seed_data.sql per lo schema v5.

Normalizzazioni applicate (concordate esplicitamente, vedi conversazione):
  - "BLACK_MASS" (usato come input in db_sottoprocessi) normalizzato in "BM"
    (il codice usato ovunque altrove: db_intermedi, db_processi.intermedio_output).
  - Righe con campi chiave mancanti (NaN) scartate: sono scarti del foglio
    Excel (es. NdO/DyO/PrO/TbO in "prodotti", mai referenziate altrove;
    righe duplicate in "macro" per MAGNETI_PERMANENTI).
"""
import openpyxl

SRC = "/mnt/user-data/uploads/recupero_materie_pippo.xlsx"
OUT = "/home/claude/recupero_fastapi/db/02_seed_data.sql"

BM_ALIASES = {"BLACK_MASS": "BM"}  # normalizza l'input dei sottoprocessi sullo stesso codice di db_intermedi


def esc(s):
    if s is None:
        return ""
    return str(s).replace("'", "''")


def norm_fasi(s):
    """Normalizza BLACK_MASS -> BM anche nei testi descrittivi (non solo nei codici)."""
    if s is None:
        return None
    return s.replace("BLACK_MASS", "BM")


def sqlval(v):
    if v is None:
        return "NULL"
    if isinstance(v, str):
        return f"'{esc(v)}'"
    return str(v)


def rows(ws, required_idx):
    """Righe valide: nessuno dei campi indicati (indici 0-based) e' None."""
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(row[i] is not None for i in required_idx):
            out.append(dict(zip([c.value for c in ws[1]], row)))
    return out


wb = openpyxl.load_workbook(SRC, data_only=True)
sql = ["-- Seed generato da recupero_materie_pippo.xlsx (database di riferimento fornito dall'utente)",
       "SET search_path TO recupero_materie;", ""]
counts = {}

# ---------------------------------------------------------------- impianti --
ws = wb["impianti"]
r = rows(ws, range(10))
counts["impianti"] = len(r)
sql.append(f"-- DB_IMPIANTI ({len(r)} righe)")
for d in r:
    sql.append(
        "INSERT INTO db_impianti (id_imp,tipologia,sotto_tecnologia,potenza_mw,anno_installazione,"
        "anno_dismissione,regione,provincia,lat,lon) VALUES "
        f"({sqlval(d['id_imp'])},{sqlval(d['tipologia'])},{sqlval(d['sotto_tecnologia'])},{d['potenza_mw']},"
        f"{d['anno_installazione']},{d['anno_dismissione']},{sqlval(d['regione'])},{sqlval(d['provincia'])},"
        f"{d['lat']},{d['lon']});"
    )

# ---------------------------------------------------------------- prodotti --
ws = wb["prodotti"]
required = ["prodotto", "nome", "categoria", "Peso_molare", "prezzo_attuale_eur_kg",
            "proiezione_prezzo_p1", "proiezione_prezzo_p2", "elemento_confronto", "parametro_equivalenza"]
headers = [c.value for c in ws[1]]
idx = [headers.index(h) for h in required]
r = rows(ws, idx)
by_code = {}
for d in r:
    by_code[d["prodotto"]] = d  # de-duplica per codice (righe ripetute identiche nel foglio)
r = list(by_code.values())
elementi = [d for d in r if d["elemento_confronto"] == d["prodotto"]]
composti = [d for d in r if d["elemento_confronto"] != d["prodotto"]]
counts["prodotti"] = len(elementi) + len(composti)
sql.append(f"\n-- DB_PRODOTTI ({len(elementi)+len(composti)} righe) — prima gli elementi puri")
sql.append("-- (elemento_confronto = se stessi), poi i composti (FK verso l'elemento)")
for d in elementi + composti:
    sql.append(
        "INSERT INTO db_prodotti (prodotto,nome,categoria,peso_molare,elemento_confronto,parametro_equivalenza,"
        "prezzo_attuale_eur_kg,proiezione_prezzo_p1,proiezione_prezzo_p2) VALUES "
        f"({sqlval(d['prodotto'])},{sqlval(d['nome'])},{sqlval(d['categoria'])},{d['Peso_molare']},"
        f"{sqlval(d['elemento_confronto'])},{d['parametro_equivalenza']},{d['prezzo_attuale_eur_kg']},"
        f"{d['proiezione_prezzo_p1']},{d['proiezione_prezzo_p2']});"
    )

# -------------------------------------------------------------------- macro --
ws = wb["macro"]
headers = [c.value for c in ws[1]]
idx = [headers.index(h) for h in ["id_macro", "tipologia", "sotto_tecnologia", "macro_componente",
                                   "massa_kg_per_mw", "prodotti", "contenuto_min", "contenuto_max"]]
r = rows(ws, idx)
counts["macro"] = len(r)
sql.append(f"\n-- DB_MACRO ({len(r)} righe: contenuto elementare per macrocomponente)")
for d in r:
    sql.append(
        "INSERT INTO db_macro (id_macro,tipologia,sotto_tecnologia,macro_componente,massa_kg_per_mw,"
        "prodotto,pct_contenuto_min,pct_contenuto_max) VALUES "
        f"({sqlval(d['id_macro'])},{sqlval(d['tipologia'])},{sqlval(d['sotto_tecnologia'])},"
        f"{sqlval(d['macro_componente'])},{d['massa_kg_per_mw']},{sqlval(d['prodotti'])},"
        f"{d['contenuto_min']},{d['contenuto_max']});"
    )

# --------------------------------------------------------------- intermedi --
ws = wb["intermedio"]
headers = [c.value for c in ws[1]]
idx = [headers.index(h) for h in ["sottoprodotto", "nome", "prezzo_attuale_eur_kg",
                                   "proiezione_prezzo_p1", "proiezione_prezzo_p2"]]
r = rows(ws, idx)
counts["intermedi"] = len(r)
sql.append(f"\n-- DB_INTERMEDI ({len(r)} righe)")
for d in r:
    sql.append(
        "INSERT INTO db_intermedi (intermedio,nome,prezzo_attuale_eur_kg,proiezione_prezzo_p1,proiezione_prezzo_p2) VALUES "
        f"({sqlval(d['sottoprodotto'])},{sqlval(d['nome'])},{d['prezzo_attuale_eur_kg']},"
        f"{d['proiezione_prezzo_p1']},{d['proiezione_prezzo_p2']});"
    )

# ---------------------------------------------------------- macro_processi --
ws = wb["macro_processi"]
headers = [c.value for c in ws[1]]
req = ["id_processo_macro", "tipologia", "sotto_tecnologia", "macro_componente", "tecnica",
       "capex_eur_ton", "opex_eur_ton", "prezzo_vendita_eur_kg", "proiezione_prezzo_p1",
       "proiezione_prezzo_p2"]
idx = [headers.index(h) for h in req]
r = rows(ws, idx)
counts["macro_processi"] = len(r)
sql.append(f"\n-- DB_MACRO_PROCESSI ({len(r)} righe)")
for d in r:
    sql.append(
        "INSERT INTO db_macro_processi (id_processo_macro,tipologia,sotto_tecnologia,macro_componente,tecnica,"
        "capex_eur_ton,opex_eur_ton,prezzo_vendita_eur_kg,proiezione_prezzo_p1,proiezione_prezzo_p2,descrizione_fasi) VALUES "
        f"({sqlval(d['id_processo_macro'])},{sqlval(d['tipologia'])},{sqlval(d['sotto_tecnologia'])},"
        f"{sqlval(d['macro_componente'])},{sqlval(d['tecnica'])},{d['capex_eur_ton']},{d['opex_eur_ton']},"
        f"{d['prezzo_vendita_eur_kg']},{d['proiezione_prezzo_p1']},{d['proiezione_prezzo_p2']},"
        f"{sqlval(norm_fasi(d.get('descrizione_fasi')))});"
    )

# ------------------------------------------------------------------ processi --
ws = wb["processi"]
headers = [c.value for c in ws[1]]
req = ["id_processo", "tipologia", "sotto_tecnologia", "macro_componente", "tecnica",
       "capex_eur_ton", "opex_eur_ton", "tipo_output", "recupero_pct_min", "recupero_pct_max"]
idx = [headers.index(h) for h in req]
r = rows(ws, idx)
counts["processi"] = len(r)
sql.append(f"\n-- DB_PROCESSI ({len(r)} righe)")
for d in r:
    tipo = "intermedio" if d["tipo_output"] == "sottoprodotto" else d["tipo_output"]
    prodotto = d.get("prodotto")
    intermedio_out = BM_ALIASES.get(d.get("sottoprodotto_output"), d.get("sottoprodotto_output"))
    sql.append(
        "INSERT INTO db_processi (id_processo,tipologia,sotto_tecnologia,macro_componente,tecnica,"
        "capex_eur_ton,opex_eur_ton,tipo_output,prodotto,intermedio_output,recupero_pct_min,recupero_pct_max,descrizione_fasi) VALUES "
        f"({sqlval(d['id_processo'])},{sqlval(d['tipologia'])},{sqlval(d['sotto_tecnologia'])},"
        f"{sqlval(d['macro_componente'])},{sqlval(d['tecnica'])},{d['capex_eur_ton']},{d['opex_eur_ton']},"
        f"{sqlval(tipo)},{sqlval(prodotto)},{sqlval(intermedio_out)},{d['recupero_pct_min']},{d['recupero_pct_max']},"
        f"{sqlval(norm_fasi(d.get('descrizione_fasi')))});"
    )

# -------------------------------------------------------------- sottoprocessi --
# Le rese originali del foglio per questi sottoprocessi erano una copia
# letterale di quelle usate a livello macrocomponente (stesso numero per
# CoSO4/NiSO4/MnSO4/LiOH sia da CELLE che da BM che da MHP) — non avevano
# senso come EFFICIENZE DI ESTRAZIONE dal contenuto elementare di BM/MHP
# (v. sotto), quindi le sostituisco con valori di tentativo plausibili per
# una raffinazione idrometallurgica (BM meno puro di MHP, quindi efficienze
# leggermente piu' basse). Il resto (capex/opex/tecnica/fasi) resta invariato.
SOTTOPROCESSI_RESA_OVERRIDE = {
    ("P_BESSN_CELLE_P3", "CoSO4"): (0.88, 0.92),
    ("P_BESSN_CELLE_P3", "NiSO4"): (0.88, 0.92),
    ("P_BESSN_CELLE_P3", "MnSO4"): (0.82, 0.88),
    ("P_BESSN_CELLE_P3", "LiOH"): (0.85, 0.90),
    ("P_BESSN_CELLE_P4", "MHP"): (0.40, 0.45),       # resa diretta (intermedio, non ha formula chimica)
    ("P_BESSN_CELLE_P4", "Li2CO3"): (0.78, 0.85),
    ("P_BESSN_CELLE_P5", "Li2CO3"): (0.80, 0.86),
    ("P_BESSN_CELLE_P5", "cake_metallico"): (0.55, 0.60),  # resa diretta
    ("P_BESSN_CELLE_P6", "CoSO4"): (0.90, 0.94),
    ("P_BESSN_CELLE_P6", "NiSO4"): (0.90, 0.94),
    ("P_BESSN_CELLE_P6", "MnSO4"): (0.85, 0.90),
    ("P_BESSN_CELLE_P6", "LiOH"): (0.85, 0.90),
}

ws = wb["sottoprocessi"]
headers = [c.value for c in ws[1]]
req = ["id_sottoprocesso", "sottoprodotto", "tecnica", "capex_eur_ton", "opex_eur_ton",
       "tipo_output", "recupero_pct_min", "recupero_pct_max"]
idx = [headers.index(h) for h in req]
r = rows(ws, idx)
counts["sottoprocessi"] = len(r)
sql.append(f"\n-- DB_SOTTOPROCESSI ({len(r)} righe) — input normalizzato su BM (era BLACK_MASS nel file")
sql.append("-- originale); rese di CoSO4/NiSO4/MnSO4/LiOH/MHP/cake_metallico sostituite (vedi commento sopra)")
for d in r:
    tipo = "intermedio" if d["tipo_output"] == "sottoprodotto" else d["tipo_output"]
    prodotto = d.get("prodotto")
    intermedio_in = BM_ALIASES.get(d["sottoprodotto"], d["sottoprodotto"])
    intermedio_out = BM_ALIASES.get(d.get("sottoprodotto_output"), d.get("sottoprodotto_output"))
    output_code = prodotto or intermedio_out
    resa_min, resa_max = SOTTOPROCESSI_RESA_OVERRIDE.get(
        (d["id_sottoprocesso"], output_code), (d["recupero_pct_min"], d["recupero_pct_max"]))
    sql.append(
        "INSERT INTO db_sottoprocessi (id_sottoprocesso,intermedio,tecnica,capex_eur_ton,opex_eur_ton,"
        "tipo_output,prodotto,intermedio_output,recupero_pct_min,recupero_pct_max,descrizione_fasi) VALUES "
        f"({sqlval(d['id_sottoprocesso'])},{sqlval(intermedio_in)},{sqlval(d['tecnica'])},{d['capex_eur_ton']},"
        f"{d['opex_eur_ton']},{sqlval(tipo)},{sqlval(prodotto)},{sqlval(intermedio_out)},"
        f"{resa_min},{resa_max},{sqlval(norm_fasi(d.get('descrizione_fasi')))});"
    )

# ------------------------------------------------- intermedi_composizione --
# Contenuto elementare di BM e MHP: non esisteva nel file originale (serviva
# perche' la conversione stechiometrica valga sempre, anche sui sottoprocessi
# che partono da un intermedio). Valori di tentativo, coerenti con una black
# mass NMC concentrata (piu' ricca di Co/Ni/Mn/Li della cella di partenza) e
# un MHP ulteriormente concentrato in Co/Ni (con tracce residue di Mn/Li).
INTERMEDI_COMPOSIZIONE = [
    ("BM", "Co", 0.12, 0.16),
    ("BM", "Ni", 0.28, 0.32),
    ("BM", "Mn", 0.10, 0.14),
    ("BM", "Li", 0.05, 0.07),
    ("MHP", "Co", 0.20, 0.24),
    ("MHP", "Ni", 0.42, 0.46),
    ("MHP", "Mn", 0.03, 0.05),
    ("MHP", "Li", 0.01, 0.02),
]
sql.append(f"\n-- DB_INTERMEDI_COMPOSIZIONE ({len(INTERMEDI_COMPOSIZIONE)} righe) — non presente nel file originale,")
sql.append("-- aggiunta perche' la stechiometria valga sempre anche a partire da un intermedio")
for intermedio, prodotto, lo, hi in INTERMEDI_COMPOSIZIONE:
    sql.append(
        f"INSERT INTO db_intermedi_composizione (intermedio,prodotto,pct_contenuto_min,pct_contenuto_max) VALUES "
        f"({sqlval(intermedio)},{sqlval(prodotto)},{lo},{hi});"
    )
counts["intermedi_composizione"] = len(INTERMEDI_COMPOSIZIONE)

# ---------------------------------------------------------- indici_rivalutazione --
ws = wb["indici_rivalutazione"]
headers = [c.value for c in ws[1]]
idx = [headers.index(h) for h in ["id_indice", "nome", "tasso_annuo"]]
r = rows(ws, idx)
counts["indici"] = len(r)
sql.append(f"\n-- DB_INDICI_RIVALUTAZIONE ({len(r)} righe)")
for d in r:
    sql.append(f"INSERT INTO db_indici_rivalutazione VALUES ({sqlval(d['id_indice'])},{sqlval(d['nome'])},{d['tasso_annuo']});")

with open(OUT, "w") as f:
    f.write("\n".join(sql) + "\n")

print(counts)
print("scritto:", OUT)
