"""Export/import Excel per le 8 tabelle. Un foglio per tabella, stesso nome
delle chiavi usate nelle route REST (routers/tables.py URL_NAMES) cosi' il
frontend puo' riusare la stessa mappatura.

Export: legge da Postgres con crud.list_rows, scrive un .xlsx con openpyxl.
Import: legge un .xlsx, valida ogni riga con i modelli Pydantic di schemas.py,
poi fa un upsert (crea se la chiave non esiste, aggiorna se esiste) tabella
per tabella, in un ordine che rispetta le chiavi esterne (db_prodotti e
db_intermedi prima di chi le referenzia).
"""
import io
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

import crud
from schemas import Impianto, Prodotto, Macro, Processo, MacroProcesso, IndiceRivalutazione, Intermedio, Sottoprocesso, IntermedioComposizione

SHEET_NAMES = {
    "db_impianti": "impianti",
    "db_prodotti": "prodotti",
    "db_macro": "macro",
    "db_intermedi_composizione": "intermedi_composizione",
    "db_macro_processi": "macro_processi",
    "db_processi": "processi",
    "db_indici_rivalutazione": "indici_rivalutazione",
    "db_intermedi": "intermedi",
    "db_sottoprocessi": "sottoprocessi",
}
TABLE_BY_SHEET = {v: k for k, v in SHEET_NAMES.items()}

MODELS = {
    "db_impianti": Impianto,
    "db_prodotti": Prodotto,
    "db_macro": Macro,
    "db_intermedi_composizione": IntermedioComposizione,
    "db_processi": Processo,
    "db_macro_processi": MacroProcesso,
    "db_indici_rivalutazione": IndiceRivalutazione,
    "db_intermedi": Intermedio,
    "db_sottoprocessi": Sottoprocesso,
}

# rispetta le FK: db_prodotti e db_intermedi prima di chi le referenzia
IMPORT_ORDER = ["db_prodotti", "db_intermedi", "db_impianti", "db_indici_rivalutazione",
                "db_macro_processi", "db_macro", "db_intermedi_composizione", "db_processi", "db_sottoprocessi"]

PERCENT_COLUMNS = {
    "proiezione_prezzo_p1", "proiezione_prezzo_p2",
    "pct_contenuto_min", "pct_contenuto_max",
    "recupero_pct_min", "recupero_pct_max",
    "tasso_annuo",
}

HEADER_FILL = PatternFill("solid", fgColor="1F5A43")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def export_workbook(tables: list[str] | None = None) -> bytes:
    tables = tables or list(SHEET_NAMES.keys())
    wb = Workbook()
    wb.remove(wb.active)
    for table in tables:
        cfg = crud.TABLES[table]
        ws = wb.create_sheet(SHEET_NAMES[table])
        cols = cfg["columns"]
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        rows = crud.list_rows(table)
        for row in rows:
            ws.append([row[c] for c in cols])
        for c_idx, col in enumerate(cols, start=1):
            if col in PERCENT_COLUMNS:
                for r_idx in range(2, len(rows) + 2):
                    ws.cell(row=r_idx, column=c_idx).number_format = "0.00%"
            width = max(12, min(40, len(col) + 4))
            ws.column_dimensions[ws.cell(row=1, column=c_idx).column_letter].width = width
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row_from_sheet(ws, header: list[str]) -> list[dict]:
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        records.append({h: v for h, v in zip(header, row)})
    return records


def import_workbook(file_bytes: bytes, dry_run: bool = False, replace: bool = False) -> dict:
    """replace=True: cancella TUTTE le righe di TUTTE le tabelle prima di
    importare (usato solo se dry_run=False: in anteprima non si cancella
    mai nulla, si simula soltanto cosa accadrebbe)."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    report = {}

    sheets_by_table = {}
    for sheet_name in wb.sheetnames:
        table = TABLE_BY_SHEET.get(sheet_name)
        if table:
            ws = wb[sheet_name]
            header = [c.value for c in ws[1]]
            sheets_by_table[table] = _row_from_sheet(ws, header)
        else:
            report.setdefault("_fogli_ignorati", []).append(sheet_name)

    if replace and not dry_run:
        report["_cancellati"] = crud.wipe_all_tables()

    for table in IMPORT_ORDER:
        if table not in sheets_by_table:
            continue
        cfg = crud.TABLES[table]
        model = MODELS[table]
        created = updated = 0
        errors = []
        for i, raw in enumerate(sheets_by_table[table], start=2):
            try:
                clean = {k: v for k, v in raw.items() if k in cfg["columns"]}
                validated = model(**clean)
                data = validated.model_dump()
                # chiave auto-generata (es. "id" seriale) lasciata vuota nel
                # foglio: nessuna riga esistente da cercare, si crea sempre.
                # In modalita' "sostituisci tutto" ogni riga e' comunque una
                # creazione (la cancellazione avviene prima, o e' simulata
                # in anteprima): non ha senso cercare righe esistenti.
                if replace or any(data.get(k) is None for k in cfg["pk"]):
                    existing = None
                else:
                    pk_values = [str(data[k]) for k in cfg["pk"]]
                    existing = crud.get_row(table, pk_values)
                if dry_run:
                    if existing:
                        updated += 1
                    else:
                        created += 1
                    continue
                if existing:
                    crud.update_row(table, pk_values, data)
                    updated += 1
                else:
                    crud.create_row(table, data)
                    created += 1
            except Exception as e:  # noqa: BLE001
                errors.append(f"riga {i}: {e}")
        report[table] = {"creati": created, "aggiornati": updated, "errori": errors,
                          "totale_foglio": len(sheets_by_table[table])}

    return report


def export_filename() -> str:
    return f"recupero_materie_export_{datetime.now():%Y%m%d_%H%M}.xlsx"
